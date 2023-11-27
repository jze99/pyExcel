from openpyxl import load_workbook
from flet import dropdown

pathFiles=[]
wb=None

def addPath(path:str, drp):
    drp.options.clear()
    pathFiles.append(path)
    wb = load_workbook(path)
    print(wb.sheetnames)
    for element in wb.sheetnames:
        drp.options.append(dropdown.Option(element))
    drp.value = wb.sheetnames[0]



    