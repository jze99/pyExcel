from CadObject import GeodesyObject, NewGeodesyObject
import openpyxl
from flet import Dropdown, dropdown

def readGeodesy(pathDirectori:str , _sheet:str):
    try:
        if pathDirectori =="":
                return False
        workbook = openpyxl.load_workbook(pathDirectori)
        sheet = workbook[_sheet]
        
        for cell in sheet[1]:
            if cell.value == 'x' or cell.value == 'X' or cell.value == 'х' or cell.value == 'Х':
                xColumn = cell.column
            if cell.value == 'y' or cell.value == 'Y' or cell.value == 'у' or cell.value == 'У':
                yColumn = cell.column
            if cell.value == 'погрешность' or cell.value == 'Погрешность':
                errorRateColumn = cell.column
            if cell.value == 'метод определения' or cell.value == 'Метод определения':
                methodDeterminationColumn = cell.column
            if cell.value == 'ИСТОЧНИК' or cell.value == 'Источник' or cell.value == 'источник':
                sourceColumn = cell.column
        for row in range(2,sheet.max_row+1):
            NewGeodesyObject.append(
                GeodesyObject(
                    x=sheet.cell(row=row, column=xColumn).value,
                    y=sheet.cell(row=row, column=yColumn).value,
                    errorRate=sheet.cell(row=row, column=errorRateColumn).value,
                    methodDetermination=sheet.cell(row=row, column=methodDeterminationColumn).value,
                    source=sheet.cell(row=row, column=sourceColumn).value
                )
            )
        workbook.close()
        return True
    except UnboundLocalError:
        return False
           
def readSheets(pathDirectori:str):
    workbook = openpyxl.load_workbook(pathDirectori)   
    sheets = workbook.sheetnames
    workbook.close()
    return sheets

