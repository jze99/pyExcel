from CadObject import NewSuperObject
import openpyxl
from openpyxl.styles import Font

def Discharge(path:str, _sheet:str):
    
    workbook = openpyxl.load_workbook(path)
    if _sheet == "Новый лист":
        sheet = workbook.create_sheet("Новый лист")
    else:
        sheet = workbook[_sheet]
        
    row = sheet.max_row+1
        
    for spObj in NewSuperObject:
        sheet.cell(row=row, column=1).value = spObj.CadNumber 
        sheet.cell(row=row, column=2).value = spObj.view
        row+=1
        sheet.cell(row=row, column=1).value = "Площадь КПТ"
        sheet.cell(row=row, column=2).value = spObj.squareKPT
        row+=1
        sheet.cell(row=row, column=1).value = "Площадь по координатам"
        sheet.cell(row=row, column=2).value = spObj.squareMIF
        row+=1
        sheet.cell(row=row, column=1).value = "% отклонения"
        sheet.cell(row=row, column=2).value = spObj.deviations
        if spObj.deviations>0:
            sheet.cell(row=row, column=3).value = "увеличение"
        elif spObj.deviations<0:
            sheet.cell(row=row, column=3).value = "уменьшение"
        row+=1
        
        for i, xSP in enumerate(spObj.x):
            sheet.cell(row=row, column=2).value = xSP
            sheet.cell(row=row, column=3).value = spObj.y[i]
            sheet.cell(row=row, column=4).value = spObj.errorRate[i]
            sheet.cell(row=row, column=5).value = spObj.methodDetermination[i]
            sheet.cell(row=row, column=6).value = spObj.sourse[i]
            if str(spObj.sourse[i]) != "ЕГРН":
                sheet.cell(row=row, column=7).value = "новая"
                for number in range(2,8):
                    sheet.cell(row=row, column=number).font = Font(bold=True)
                    
            row+=1
        row+=1
    workbook.save(path)
    workbook.close()
            
            