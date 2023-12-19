import openpyxl
from CadObject import KPTObject, NewKPTCadObject

def readKPT(pathDirectori:str):
    try:
        if pathDirectori =="":
            return False
        workbook = openpyxl.load_workbook(pathDirectori)
        sheet = workbook['2']
        for row in range(2, sheet.max_row+1):
            NewKPTCadObject.append(KPTObject(CadNumber=sheet.cell(row=row, column=1).value, view=sheet.cell(row=row, column=2).value, square=sheet.cell(row=row, column=3).value))
        sheet = workbook['1']
        cd=""
        x=[]
        y=[]
        errorRate=[]
        methodDetermination=[]
        source=[]
        i=0
        for row in range(2,sheet.max_row+1): 
            if cd == "":
                cd=NewKPTCadObject[i].CadNumber
            if sheet.cell(row=row,column=1).value != cd:
                if not x:
                    i+=1
                    cd=NewKPTCadObject[i].CadNumber
                    if sheet.cell(column=1,row=row).value == cd:
                        x.append(sheet.cell(row=row, column=3).value) 
                        y.append(sheet.cell(row=row, column=4).value) 
                        errorRate.append(sheet.cell(row=row, column=5).value)
                        methodDetermination.append(sheet.cell(row=row, column=6).value)
                        source.append(sheet.cell(row=row, column=7).value)
                else:
                    NewKPTCadObject[i].appedt(
                        _x=x,
                        _y=y,
                        _errorRate=errorRate,
                        _methodDetermination=methodDetermination,
                        _source=source
                    )
                    x.clear()
                    y.clear()
                    errorRate.clear()
                    methodDetermination.clear()
                    source.clear()
                    i+=1
                    cd=NewKPTCadObject[i].CadNumber
            if sheet.cell(column=1,row=row).value == cd:
                x.append(sheet.cell(row=row, column=3).value) 
                y.append(sheet.cell(row=row, column=4).value) 
                errorRate.append(sheet.cell(row=row, column=5).value)
                methodDetermination.append(sheet.cell(row=row, column=6).value)
                source.append(sheet.cell(row=row, column=7).value)
        workbook.close()
        return True
    except KeyError:
        return False
