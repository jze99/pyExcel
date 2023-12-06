from pars import newObjets
import openpyxl

def uploadPath(file_path: str):
    
    if file_path == "":
        exit()
    
    workbook = openpyxl.load_workbook(file_path)
    
    
    if '1' not in workbook.sheetnames:
        workbook.create_sheet('1')
        sheet = workbook['1']
    else:
        sheet = workbook['1']
    
    row_number = sheet.max_row+1
    
    for number in newObjets:
        if number.x:
            for ind in range(len(number.x)):
                sheet['A' + str(row_number+ind)] = number.cadNumber
                sheet['B' + str(row_number+ind)] = number.view
                sheet['C' + str(row_number+ind)] = number.x[ind]
                sheet['D' + str(row_number+ind)] = number.y[ind]
            if number.errorRate:
                for ind1 in range(len(number.errorRate)):
                    sheet['E' + str(row_number+ind1)] = number.errorRate[ind1] 
            row_number+=ind       
            row_number += 1
    
    if '2' not in workbook.sheetnames:
        workbook.create_sheet('2')
        sheet = workbook['2']
    else:
        sheet = workbook['2']
        
    row_number = sheet.max_row+1
    
    for number in newObjets:
        sheet['A' + str(row_number)] = number.cadNumber
        sheet['B' + str(row_number)] = number.view
        sheet['C' + str(row_number)] = number.square
        row_number += 1

    workbook.save(file_path)
